# exporter.py
# ---------------------------------------------------------------
# Exports cleaned data to CSV and formatted Excel
# Input:  data/internships_clean.csv
# Output: data/internships_export.csv
#         data/internships_export.xlsx
# ---------------------------------------------------------------

import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter


# ---------------------------------------------------------------
# STEP 1: LOAD CLEAN DATA
# ---------------------------------------------------------------

def load_clean_data(path='data/internships_clean.csv'):
    """Load the cleaned CSV produced by cleaner.py"""
    if not os.path.exists(path):
        print(f"❌ File not found: {path}")
        print("   Run cleaner.py first!")
        return None

    df = pd.read_csv(path)
    print(f"✅ Loaded {len(df)} rows from {path}")
    return df


# ---------------------------------------------------------------
# STEP 2: EXPORT TO CSV
# ---------------------------------------------------------------

def export_csv(df, path='data/internships_export.csv'):
    """
    Export to CSV with UTF-8 BOM encoding.
    utf-8-sig ensures Excel opens it correctly — without this,
    Indian characters like ₹ show up as garbage in Excel.
    """
    os.makedirs('data', exist_ok=True)
    df.to_csv(path, index=False, encoding='utf-8-sig')
    print(f"✅ CSV exported: {path}  ({len(df)} rows)")
    return path


# ---------------------------------------------------------------
# STEP 3: EXPORT TO FORMATTED EXCEL
# ---------------------------------------------------------------

def export_excel(df, path='data/internships_export.xlsx'):
    """
    Exports DataFrame to a professionally formatted Excel file.
    Uses XlsxWriter engine directly — more reliable than openpyxl post-formatting.
    """
    os.makedirs('data', exist_ok=True)

    # Use XlsxWriter engine — writes formatting IN ONE PASS (no reopen needed)
    writer = pd.ExcelWriter(path, engine='xlsxwriter')

    # Write main data sheet
    df.to_excel(writer, index=False, sheet_name='Internships')

    # Get xlsxwriter objects to apply formatting
    workbook  = writer.book
    worksheet = writer.sheets['Internships']

    # ---------------------------------------------------------------
    # DEFINE FORMATS
    # ---------------------------------------------------------------

    # Header format: dark navy + white bold text + centered
    header_fmt = workbook.add_format({
        'bold':       True,
        'font_name':  'Calibri',
        'font_size':  11,
        'font_color': '#FFFFFF',
        'bg_color':   '#1F3864',
        'align':      'center',
        'valign':     'vcenter',
        'border':     1,
        'border_color': '#BFBFBF',
        'text_wrap':  True,
    })

    # Even row format: light blue background
    even_fmt = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'bg_color':   '#DCE6F1',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
    })

    # Odd row format: white background
    odd_fmt = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'bg_color':   '#FFFFFF',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
    })

    # Stipend_Avg column: light green
    stipend_fmt = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'bg_color':   '#E2EFDA',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
        'num_format': '₹#,##0',
    })

    # Hyperlink format: blue underlined
    link_fmt = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'font_color': '#0563C1',
        'underline':  True,
        'bg_color':   '#DCE6F1',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
    })

    link_fmt_odd = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'font_color': '#0563C1',
        'underline':  True,
        'bg_color':   '#FFFFFF',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
    })

    # Date format for Scraped At column — fixes the ##### issue
    date_fmt = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'bg_color':   '#DCE6F1',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
        'num_format': 'yyyy-mm-dd hh:mm',  # explicit date format
    })

    date_fmt_odd = workbook.add_format({
        'font_name':  'Calibri',
        'font_size':  10,
        'bg_color':   '#FFFFFF',
        'border':     1,
        'border_color': '#BFBFBF',
        'valign':     'vcenter',
        'num_format': 'yyyy-mm-dd hh:mm',
    })

    # ---------------------------------------------------------------
    # COLUMN SETUP
    # ---------------------------------------------------------------

    headers = list(df.columns)

    # Column widths — set per column name
    col_widths = {
        'Title':          28,
        'Company':        28,
        'Location':       16,
        'Stipend':        22,
        'Stipend_Min':    12,
        'Stipend_Max':    12,
        'Stipend_Avg':    13,
        'Duration':       12,
        'Duration_Months':16,
        'Skills':         35,
        'Date Posted':    16,
        'Link':           16,
        'Scraped At':     18,
    }

    for col_idx, col_name in enumerate(headers):
        width = col_widths.get(col_name, 15)
        worksheet.set_column(col_idx, col_idx, width)

    # ---------------------------------------------------------------
    # WRITE HEADER ROW WITH FORMATTING
    # ---------------------------------------------------------------

    worksheet.set_row(0, 30)   # header row height = 30px

    for col_idx, col_name in enumerate(headers):
        worksheet.write(0, col_idx, col_name, header_fmt)

    # ---------------------------------------------------------------
    # WRITE DATA ROWS WITH FORMATTING
    # ---------------------------------------------------------------

    link_col_idx        = headers.index('Link')         if 'Link'        in headers else None
    stipend_avg_col_idx = headers.index('Stipend_Avg')  if 'Stipend_Avg' in headers else None
    scraped_at_col_idx  = headers.index('Scraped At')   if 'Scraped At'  in headers else None

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=1):
        # Row height
        worksheet.set_row(row_idx, 18)

        # Pick base format for this row (alternating)
        is_even  = (row_idx % 2 == 0)
        base_fmt = even_fmt if is_even else odd_fmt

        for col_idx, col_name in enumerate(headers):
            value = getattr(row_data, col_name.replace(' ', '_').replace('/', '_'))

            # Handle NaN — write blank instead
            if pd.isna(value) if not isinstance(value, str) else value == 'N/A':
                worksheet.write_blank(row_idx, col_idx, None, base_fmt)
                continue

            # Link column: write as URL with display text
            if col_idx == link_col_idx:
                fmt = link_fmt if is_even else link_fmt_odd
                try:
                    worksheet.write_url(row_idx, col_idx, value, fmt, 'View Internship')
                except Exception:
                    worksheet.write(row_idx, col_idx, value, fmt)

            # Stipend_Avg: green background + currency format
            elif col_idx == stipend_avg_col_idx:
                worksheet.write_number(row_idx, col_idx, int(value), stipend_fmt)

            # Scraped At: explicit date format to prevent #####
            elif col_idx == scraped_at_col_idx:
                fmt = date_fmt if is_even else date_fmt_odd
                worksheet.write(row_idx, col_idx, str(value), fmt)

            # Everything else: normal write
            else:
                worksheet.write(row_idx, col_idx, value, base_fmt)

    # ---------------------------------------------------------------
    # FREEZE HEADER + AUTOFILTER
    # ---------------------------------------------------------------

    worksheet.freeze_panes(1, 0)       # freeze row 1
    worksheet.autofilter(0, 0, len(df), len(headers) - 1)

    # ---------------------------------------------------------------
    # SUMMARY SHEET
    # ---------------------------------------------------------------

    ws2 = workbook.add_worksheet('Summary')

    title_fmt   = workbook.add_format({'bold': True, 'font_size': 14, 'font_color': '#1F3864', 'font_name': 'Calibri'})
    section_fmt = workbook.add_format({'bold': True, 'font_size': 11, 'bg_color': '#DCE6F1', 'font_name': 'Calibri', 'border': 1})
    label_fmt2  = workbook.add_format({'bold': True, 'font_size': 11, 'font_name': 'Calibri'})
    value_fmt2  = workbook.add_format({'font_size': 11, 'font_name': 'Calibri'})
    italic_fmt  = workbook.add_format({'italic': True, 'font_size': 10, 'font_color': '#808080', 'font_name': 'Calibri'})

    ws2.set_column(0, 0, 35)
    ws2.set_column(1, 1, 25)

    ws2.write(0, 0, 'JobScrapeX — Internship Data Summary', title_fmt)
    ws2.write(1, 0, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', italic_fmt)

    rows = [
        (None, 'OVERVIEW', ''),
        (None, 'Total Internships',    len(df)),
        (None, 'Unique Companies',     int(df['Company'].nunique())),
        (None, 'Unique Locations',     int(df['Location'].nunique())),
        (None, 'Work from Home',       int((df['Location'] == 'Work from Home').sum())),
        (None, '', ''),
        (None, 'STIPEND STATS', ''),
        (None, 'Average Stipend',      f"₹{df['Stipend_Avg'].mean():,.0f}/month" if 'Stipend_Avg' in df else 'N/A'),
        (None, 'Highest Stipend',      f"₹{int(df['Stipend_Max'].max()):,}/month" if 'Stipend_Max' in df else 'N/A'),
        (None, 'Lowest Stipend',       f"₹{int(df['Stipend_Min'].min()):,}/month" if 'Stipend_Min' in df else 'N/A'),
        (None, '', ''),
        (None, 'TOP 5 LOCATIONS', ''),
    ]

    for loc, cnt in df['Location'].value_counts().head(5).items():
        rows.append((None, f'  {loc}', f'{cnt} internships'))

    rows.append((None, '', ''))
    rows.append((None, 'DURATION BREAKDOWN', ''))
    for dur, cnt in df['Duration'].value_counts().items():
        rows.append((None, f'  {dur}', f'{cnt} internships'))

    section_labels = {'OVERVIEW', 'STIPEND STATS', 'TOP 5 LOCATIONS', 'DURATION BREAKDOWN'}

    for i, (_, label, value) in enumerate(rows, start=3):
        if label in section_labels:
            ws2.write(i, 0, label, section_fmt)
            ws2.write(i, 1, '',    section_fmt)
        elif label == '':
            ws2.write(i, 0, '', value_fmt2)
        else:
            ws2.write(i, 0, label, label_fmt2)
            ws2.write(i, 1, value, value_fmt2)

    # ---------------------------------------------------------------
    # SAVE
    # ---------------------------------------------------------------
    writer.close()
    print(f"✅ Excel exported: {path}")
    print(f"   Sheets: 'Internships' (data) + 'Summary' (stats)")
    return path


# ---------------------------------------------------------------
# STEP 4: EXPORT FILTERED SUBSETS
# ---------------------------------------------------------------

def export_filtered(df, keyword=None, location=None,
                    min_stipend=0, path_suffix='filtered'):
    """
    Export a filtered subset to both CSV and Excel.
    Imports filter functions from cleaner.py.
    """
    from cleaner import apply_all_filters

    filtered = apply_all_filters(
        df,
        keyword=keyword,
        location=location,
        min_stipend=min_stipend
    )

    if filtered.empty:
        print(f"⚠️  No results for filter — nothing exported.")
        return

    csv_path   = f'data/internships_{path_suffix}.csv'
    excel_path = f'data/internships_{path_suffix}.xlsx'

    export_csv(filtered, csv_path)
    export_excel(filtered, excel_path)
    return filtered


# ---------------------------------------------------------------
# RUN
# ---------------------------------------------------------------

if __name__ == '__main__':
    print("=" * 55)
    print("  JobScrapeX — Exporter")
    print("=" * 55)

    # Load clean data
    df = load_clean_data()
    if df is None:
        exit()

    # Export full dataset
    print("\n--- Exporting Full Dataset ---")
    export_csv(df)
    export_excel(df)

    # Export filtered subset examples
    print("\n--- Exporting Filtered Subsets ---")

    # All Ahmedabad internships
    export_filtered(df,
        location='ahmedabad',
        path_suffix='ahmedabad'
    )

    # High paying internships (₹15,000+)
    export_filtered(df,
        min_stipend=15000,
        path_suffix='high_paying'
    )

    print("\n✅ All exports complete!")
    print("   Check your /data folder — open the .xlsx files in Excel")